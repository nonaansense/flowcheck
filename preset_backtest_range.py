"""
preset_backtest_range.py — Multi-day Bullflow preset backtest → Excel.

Iterates every trading weekday in a date range, runs the same per-day
preset backtest logic (preset_backtest.collect_day), and writes results
to an .xlsx with ONE TAB PER DATE, delivered to Telegram as a document.

Weekends are skipped automatically (the options market is closed, so a
weekend replay returns an empty stream). Range is capped at 31 calendar
days per run (a full month).

Each date tab has one row per qualifying alert with columns:
  Time, Preset Type, Ticker, Direction, Strike, Expiry, DTE, Moneyness,
  Total Premium, Contracts, Trade Price, Entry, Trail Offset, Stock @ Alert,
  Earnings, Sweep

A leading Summary tab lists per-date counts + total premium.

Telegram command: /preset_backtest_range YYYY-MM-DD YYYY-MM-DD
"""
import os, re, time, threading
from datetime import datetime, timedelta

MAX_RANGE_DAYS = 31   # a full month of calendar days per run


def _weekdays_in_range(start: str, end: str) -> list:
    """Return YYYY-MM-DD strings for each Mon-Fri in [start, end] inclusive."""
    d0 = datetime.strptime(start, "%Y-%m-%d").date()
    d1 = datetime.strptime(end,   "%Y-%m-%d").date()
    out = []
    cur = d0
    while cur <= d1:
        if cur.weekday() < 5:   # 0=Mon .. 4=Fri
            out.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    return out


# Column layout for each per-date sheet
_COLUMNS = [
    ("Time",          lambda a: a.get("time_str", "")),
    ("Preset Type",   lambda a: a.get("preset_type", "")),
    ("Ticker",        lambda a: a.get("ticker", "")),
    ("Direction",     lambda a: a.get("direction", "").upper()),
    ("Strike",        lambda a: a.get("strike", "")),
    ("Expiry",        lambda a: a.get("expiry", "")),
    ("DTE",           lambda a: a.get("dte", 0)),
    ("Moneyness",     lambda a: a.get("moneyness", "")),
    ("Total Premium", lambda a: round(a.get("premium", 0), 2)),
    ("Contracts",     lambda a: a.get("contracts", 0)),
    ("Trade Price",   lambda a: round(a.get("price", 0), 2)),
    ("Entry",         lambda a: round(a.get("entry_price", 0), 2)),
    ("Trail Offset",  lambda a: -round(a.get("trail_offset", 0), 2)),
    ("Stock @ Alert", lambda a: round(a.get("stock_px", 0), 2)),
    ("Earnings",      lambda a: a.get("earnings_str") or ""),
    ("Sweep",         lambda a: "YES" if a.get("sweep") else ""),
    ("30M Play",      lambda a: "REVERSAL" if a.get("playbook") == "reversal" else "FOLLOW"),
    ("EMA Fast",      lambda a: a.get("ema_fast") or ""),
    ("EMA Slow",      lambda a: a.get("ema_slow") or ""),
    ("EMA State",     lambda a: ("5>12" if (a.get("ema_fast") or 0) > (a.get("ema_slow") or 0)
                                 else "5<12") if (a.get("ema_fast") and a.get("ema_slow"))
                                else "no data"),
    ("Early <10:30",  lambda a: "YES — reversal risk" if a.get("is_early") else ""),
    ("Entry Filled",  lambda a: "YES" if (a.get("pnl") or {}).get("entry_filled") else "NO"),
    ("Leg1 Exit",     lambda a: (a.get("pnl") or {}).get("leg1_exit")
                                if (a.get("pnl") or {}).get("leg1_exit") is not None else ""),
    ("Leg1 Reason",   lambda a: (a.get("pnl") or {}).get("leg1_reason") or ""),
    ("Leg2 Exit",     lambda a: (a.get("pnl") or {}).get("leg2_exit")
                                if (a.get("pnl") or {}).get("leg2_exit") is not None else ""),
    ("Leg2 Reason",   lambda a: (a.get("pnl") or {}).get("leg2_reason") or ""),
    ("Contracts",     lambda a: (a.get("pnl") or {}).get("contracts") or ""),
    ("Capital $",     lambda a: (a.get("pnl") or {}).get("capital") or ""),
    ("P/L $",         lambda a: (a.get("pnl") or {}).get("pnl_usd_trail")
                                if (a.get("pnl") or {}).get("pnl_usd_trail") is not None else ""),
    ("P/L %",         lambda a: (a.get("pnl") or {}).get("pnl_pct_trail")
                                if (a.get("pnl") or {}).get("pnl_pct_trail") is not None else ""),
    ("P/L $ NO trail",lambda a: (a.get("pnl") or {}).get("pnl_usd_notrail")
                                if (a.get("pnl") or {}).get("pnl_usd_notrail") is not None else ""),
    ("P/L % NO trail",lambda a: (a.get("pnl") or {}).get("pnl_pct_notrail")
                                if (a.get("pnl") or {}).get("pnl_pct_notrail") is not None else ""),
    ("Trail Cost $",  lambda a: (a.get("pnl") or {}).get("trail_cost_usd")
                                if (a.get("pnl") or {}).get("trail_cost_usd") is not None else ""),
    ("Traded",        lambda a: a.get("traded") or "original"),
    ("Orig P/L $",    lambda a: (a.get("pnl_orig") or {}).get("pnl_usd_trail")
                                if (a.get("pnl_orig") or {}).get("pnl_usd_trail") is not None else ""),
    ("Orig P/L %",    lambda a: (a.get("pnl_orig") or {}).get("pnl_pct_trail")
                                if (a.get("pnl_orig") or {}).get("pnl_pct_trail") is not None else ""),
    ("Orig Peak %",   lambda a: (a.get("pnl_orig") or {}).get("mfe_pct")
                                if (a.get("pnl_orig") or {}).get("mfe_pct") is not None else ""),
    ("Roll Edge $",   lambda a: a.get("roll_edge_usd")
                                if a.get("roll_edge_usd") is not None else ""),
    ("Roll Verdict",  lambda a: ("ROLL WON" if (a.get("roll_edge_usd") or 0) > 0
                                 else "ROLL LOST" if (a.get("roll_edge_usd") or 0) < 0
                                 else "") if a.get("roll_edge_usd") is not None else ""),
    ("Roll Expiry",   lambda a: (a.get("roll") or {}).get("expiry") or ""),
    ("Roll Reason",   lambda a: (a.get("roll") or {}).get("reason") or ""),
    ("Roll Price",    lambda a: (a.get("roll") or {}).get("price") or ""),
    ("Roll Entry",    lambda a: (a.get("roll") or {}).get("entry") or ""),
    ("Roll T1",       lambda a: (a.get("roll") or {}).get("target1") or ""),
    ("Roll T2",       lambda a: (a.get("roll") or {}).get("target2") or ""),
    ("Roll Trail",    lambda a: -(a.get("roll") or {}).get("trail", 0)
                                if (a.get("roll") or {}).get("trail") else ""),
    ("Max Price",     lambda a: (a.get("pnl") or {}).get("max_price")
                                if (a.get("pnl") or {}).get("max_price") is not None else ""),
    ("MFE % (peak)",  lambda a: (a.get("pnl") or {}).get("mfe_pct")
                                if (a.get("pnl") or {}).get("mfe_pct") is not None else ""),
    ("MAE % (heat)",  lambda a: (a.get("pnl") or {}).get("mae_pct")
                                if (a.get("pnl") or {}).get("mae_pct") is not None else ""),
    ("Max DD %",      lambda a: (a.get("pnl") or {}).get("max_dd_pct")
                                if (a.get("pnl") or {}).get("max_dd_pct") is not None else ""),
    ("Days Held",     lambda a: (a.get("pnl") or {}).get("days_held") or ""),
    ("Fill Time",     lambda a: (a.get("pnl") or {}).get("fill_time") or ""),
    ("Fill Source",   lambda a: (a.get("pnl") or {}).get("fill_source") or ""),
    ("Target 1",      lambda a: a.get("target1") or ""),
    ("Target 2",      lambda a: a.get("target2") or ""),
    ("T1 Hit",        lambda a: "YES" if (a.get("pnl") or {}).get("t1_hit") else
                                ("NO" if (a.get("pnl") or {}).get("entry_filled") else "")),
    ("T2 Hit",        lambda a: "YES" if (a.get("pnl") or {}).get("t2_hit") else
                                ("NO" if (a.get("pnl") or {}).get("entry_filled") else "")),
    ("Max Profit %",  lambda a: (a.get("pnl") or {}).get("max_profit_pct")
                                if (a.get("pnl") or {}).get("max_profit_pct") is not None else ""),
    ("Max Profit $/ct", lambda a: (a.get("pnl") or {}).get("max_profit_per_contract")
                                if (a.get("pnl") or {}).get("max_profit_per_contract") is not None else ""),
]


def _safe_sheet_title(date_str: str) -> str:
    # Excel sheet titles: <=31 chars, no : \ / ? * [ ]. Date is already safe.
    return date_str[:31]


def _build_workbook(day_results: list, start: str, end: str) -> str:
    """day_results: list of collect_day() dicts (in date order).
    Returns the path to the written .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    center      = Alignment(horizontal="center")

    # ── Summary tab (first) ──
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Date", "Alerts", "Total Premium", "Preset Events",
                    "Filled", "Wins", "Losses", "Win Rate %", "Avg P/L %",
                    "P/L $", "P/L $ NO trail", "Trail Cost $",
                    "Avg MFE %", "Avg MAE %", "T1 Hit", "T1 %", "T2 Hit", "T2 %",
                    "Rolled", "Roll Edge $", "Note"])
    for c in range(1, 22):
        cell = summary.cell(row=1, column=c)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center

    all_pnls, all_mfes, all_maes = [], [], []
    all_usd_t, all_usd_nt, all_edges = [], [], []
    all_t1 = all_t2 = all_filled = 0
    for day in day_results:
        alerts = day["alerts"]
        total_prem = sum(a.get("premium", 0) for a in alerts)
        pnls = [(a.get("pnl") or {}).get("pnl_pct_trail") for a in alerts]
        pnls = [p for p in pnls if p is not None]
        usd_t  = [(a.get("pnl") or {}).get("pnl_usd_trail") for a in alerts]
        usd_t  = [u for u in usd_t if u is not None]
        usd_nt = [(a.get("pnl") or {}).get("pnl_usd_notrail") for a in alerts]
        usd_nt = [u for u in usd_nt if u is not None]
        mfes = [(a.get("pnl") or {}).get("mfe_pct") for a in alerts]
        mfes = [m for m in mfes if m is not None]
        maes = [(a.get("pnl") or {}).get("mae_pct") for a in alerts]
        maes = [m for m in maes if m is not None]
        filled = [a for a in alerts if (a.get("pnl") or {}).get("entry_filled")]
        t1s = sum(1 for a in filled if (a.get("pnl") or {}).get("t1_hit"))
        t2s = sum(1 for a in filled if (a.get("pnl") or {}).get("t2_hit"))
        all_t1 += t1s; all_t2 += t2s; all_filled += len(filled)
        all_pnls.extend(pnls); all_mfes.extend(mfes); all_maes.extend(maes)
        all_usd_t.extend(usd_t); all_usd_nt.extend(usd_nt)
        wins   = sum(1 for p in pnls if p > 0)
        losses = sum(1 for p in pnls if p <= 0)
        win_rate = round(wins / len(pnls) * 100, 1) if pnls else ""
        avg_pnl  = round(sum(pnls) / len(pnls), 1) if pnls else ""
        avg_mfe  = round(sum(mfes) / len(mfes), 1) if mfes else ""
        avg_mae  = round(sum(maes) / len(maes), 1) if maes else ""
        t1_rate  = round(t1s / len(filled) * 100, 1) if filled else ""
        t2_rate  = round(t2s / len(filled) * 100, 1) if filled else ""
        note = day.get("error", "") or ("no alerts" if not alerts else "")
        d_usd_t  = round(sum(usd_t), 2)  if usd_t  else ""
        d_usd_nt = round(sum(usd_nt), 2) if usd_nt else ""
        d_cost   = round(sum(usd_nt) - sum(usd_t), 2) if (usd_t and usd_nt) else ""
        edges = [a.get("roll_edge_usd") for a in alerts
                 if a.get("roll_edge_usd") is not None]
        all_edges.extend(edges)
        d_rolled = len(edges)
        d_edge   = round(sum(edges), 2) if edges else ""
        summary.append([day["date"], len(alerts), round(total_prem, 2),
                        day.get("preset_events", 0), len(pnls), wins, losses,
                        win_rate, avg_pnl, d_usd_t, d_usd_nt, d_cost,
                        avg_mfe, avg_mae, t1s, t1_rate, t2s, t2_rate,
                        d_rolled, d_edge, note])

    # Overall totals row
    if all_pnls:
        t_wins = sum(1 for p in all_pnls if p > 0)
        summary.append([])
        summary.append(["TOTAL", "", "", "", len(all_pnls), t_wins,
                        len(all_pnls) - t_wins,
                        round(t_wins / len(all_pnls) * 100, 1),
                        round(sum(all_pnls) / len(all_pnls), 1),
                        round(sum(all_usd_t), 2) if all_usd_t else "",
                        round(sum(all_usd_nt), 2) if all_usd_nt else "",
                        round(sum(all_usd_nt) - sum(all_usd_t), 2) if (all_usd_t and all_usd_nt) else "",
                        round(sum(all_mfes) / len(all_mfes), 1) if all_mfes else "",
                        round(sum(all_maes) / len(all_maes), 1) if all_maes else "",
                        all_t1,
                        round(all_t1 / all_filled * 100, 1) if all_filled else "",
                        all_t2,
                        round(all_t2 / all_filled * 100, 1) if all_filled else "",
                        len(all_edges),
                        round(sum(all_edges), 2) if all_edges else "",
                        "all filled trades"])
        for c in range(1, 22):
            summary.cell(row=summary.max_row, column=c).font = Font(bold=True)

    for col, width in zip("ABCDEFGHIJKLMNOPQRSTU", (12, 8, 16, 14, 8, 7, 8, 11, 11, 13, 15, 12, 11, 11, 8, 8, 8, 8, 8, 12, 24)):
        summary.column_dimensions[col].width = width

    # ── One tab per date ──
    for day in day_results:
        ws = wb.create_sheet(title=_safe_sheet_title(day["date"]))
        ws.append([c[0] for c in _COLUMNS])
        for i in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=1, column=i)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center

        # Sort alerts by time for readability
        for a in sorted(day["alerts"], key=lambda x: x.get("timestamp") or 0):
            ws.append([fn(a) for _, fn in _COLUMNS])

        # Reasonable column widths
        widths = [11, 22, 8, 10, 9, 10, 6, 10, 14, 10, 11, 9, 12, 13, 16, 7, 11, 20, 12, 11, 11, 11, 11, 13, 13, 15, 15, 13, 12, 11, 11, 13, 13, 10, 10, 11, 20, 10, 10, 8, 8, 13, 15, 10, 10, 11, 12, 10, 10, 10, 10, 16, 12, 12, 12, 12, 12, 10, 11]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
        ws.freeze_panes = "A2"

    out_dir = "/tmp"
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"preset_backtest_{start}_to_{end}.xlsx")
    wb.save(path)
    return path


def _run_range_thread(start: str, end: str, bot_token: str, chat_id: str):
    from sms import send_telegram, send_telegram_document
    from preset_backtest import collect_day
    import bullflow_presets as bp

    days = _weekdays_in_range(start, end)
    if not days:
        send_telegram(f"No weekdays in range {start} → {end}.", bot_token, chat_id)
        return

    send_telegram(
        f"📊 Preset range backtest: {start} → {end}\n"
        f"{len(days)} trading days | filters ≥{bp._fmt_prem(bp.MIN_PREMIUM)}, "
        f"≤{bp.MAX_DTE}d DTE\n"
        f"~{len(days)*6} min at 60x. I'll send the Excel when done.",
        bot_token, chat_id)

    day_results = []
    total_alerts = 0
    for i, date in enumerate(days, 1):
        try:
            day = collect_day(date)
        except Exception as e:
            day = {"date": date, "alerts": [], "preset_events": 0,
                   "event_count": 0, "seen_names": {}, "error": str(e)}
        day_results.append(day)
        total_alerts += len(day["alerts"])
        # Progress ping every 5 days
        if i % 5 == 0 and i < len(days):
            send_telegram(f"… {i}/{len(days)} days processed "
                          f"({total_alerts} alerts so far)", bot_token, chat_id)

    try:
        path = _build_workbook(day_results, start, end)
    except Exception as e:
        send_telegram(f"❌ Excel build error: {e}", bot_token, chat_id)
        return

    ok = send_telegram_document(
        path, bot_token, chat_id,
        caption=f"Preset backtest {start} → {end}: {total_alerts} alerts "
                f"across {len(days)} trading days (tab per date).")
    if not ok:
        send_telegram("❌ Could not send the Excel file to Telegram.", bot_token, chat_id)
    try:
        os.remove(path)
    except Exception:
        pass


def start_range_backtest(start: str, end: str, bot_token: str, chat_id: str):
    """
    Validate the range and launch in a background thread.
    Returns (ok: bool, message: str) — message explains any rejection.
    """
    if not (re.match(r'^\d{4}-\d{2}-\d{2}$', start) and re.match(r'^\d{4}-\d{2}-\d{2}$', end)):
        return False, "Both dates must be YYYY-MM-DD. e.g. /preset_backtest_range 2026-06-01 2026-06-30"
    try:
        d0 = datetime.strptime(start, "%Y-%m-%d").date()
        d1 = datetime.strptime(end,   "%Y-%m-%d").date()
    except Exception:
        return False, "Invalid date(s). Use YYYY-MM-DD."
    if d1 < d0:
        return False, "End date is before start date."
    span = (d1 - d0).days + 1
    if span > MAX_RANGE_DAYS:
        return False, f"Range too large ({span} days). Max is {MAX_RANGE_DAYS} days (a full month)."

    t = threading.Thread(target=_run_range_thread,
                         args=(start, end, bot_token, chat_id),
                         daemon=True, name=f"preset_range_{start}_{end}")
    t.start()
    return True, ""

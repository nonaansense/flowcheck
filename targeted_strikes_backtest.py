"""
targeted_strikes_backtest.py — Targeted strike/expiry stacking backtest runner.

Connects to Bullflow's backtesting SSE endpoint, replays a full trading day
through the targeted-strikes detection logic in isolated state (never
touches production _TARGETED), and reports results to Telegram when done.

Backtesting always tracks both calls and puts.

URL: https://api.bullflow.io/v1/streaming/backtesting?key={KEY}&date={DATE}&speed=60

Uses actual event timestamps from the stream (not real-clock + speed
scaling) for accurate same-day accumulation tracking, and the stream's
estTimestamp for the early-session (before 10:25am ET) check.

Telegram command: /targeted_backtest YYYY-MM-DD [detail]

For a date RANGE, call start_backtest_range(), which walks each trading
day sequentially (same approach as preset_backtest_range.py) and sends
one combined summary at the end.
"""
import os, re, time, json, threading, requests
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

ET    = ZoneInfo("America/New_York")
SPEED = 60   # Bullflow playback speed multiplier
MAX_RANGE_DAYS = 31   # matches preset_backtest_range.py — a full month of calendar days per run

TARGETED_FILTER  = os.environ.get("TARGETED_STRIKES_FILTER_NAME", "Targeted_Strikes_Expiry")
THRESHOLD        = int(os.environ.get("TARGETED_STRIKES_THRESHOLD", "4"))
EARLY_CUTOFF_STR = os.environ.get("TARGETED_STRIKES_EARLY_CUTOFF", "10:25")
SKIP_EARLY       = os.environ.get("TARGETED_STRIKES_SKIP_EARLY", "false").lower() in ("true","1","yes","on")
GATE_UNTIL_CUTOFF = os.environ.get("TARGETED_STRIKES_GATE_UNTIL_CUTOFF", "false").lower() in ("true","1","yes","on")


def _early_cutoff():
    try:
        hh, mm = EARLY_CUTOFF_STR.split(":")
        return int(hh), int(mm)
    except Exception:
        return 10, 25


def _is_early_str(est_str: str) -> bool:
    """est_str looks like '2026-06-01T09:42:11-04:00' or similar; fall back False."""
    try:
        hh_mm = est_str[11:16]
        hh, mm = int(hh_mm[:2]), int(hh_mm[3:5])
        cutoff_hh, cutoff_mm = _early_cutoff()
        return (hh, mm) < (cutoff_hh, cutoff_mm)
    except Exception:
        return False


def _build_url(api_key: str, date: str) -> str:
    return (f"https://api.bullflow.io/v1/streaming/backtesting"
            f"?key={api_key}&date={date}&speed={SPEED}")


def _parse_occ(symbol: str, as_of: str = None) -> dict | None:
    """
    Parse OCC option symbol: O:GOOGL260717C00370000

    `as_of` (YYYY-MM-DD) is the date to measure DTE against. In a backtest
    this MUST be the replay date — measuring against datetime.now() would
    make every historical contract look 0DTE and silently disqualify it from
    swing scoring.
    """
    if not symbol:
        return None
    m = re.search(r'O:([A-Z]+)(\d{2})(\d{2})(\d{2})([CP])(\d+)', symbol)
    if not m:
        return None
    ticker      = m.group(1)
    yy, mm, dd  = m.group(2), m.group(3), m.group(4)
    option_type = "call" if m.group(5) == "C" else "put"
    strike_raw  = int(m.group(6)) / 1000.0
    strike      = str(int(strike_raw)) if strike_raw == int(strike_raw) else f"{strike_raw:.1f}"
    expiry      = f"{mm}/{dd}/{yy}"
    try:
        exp_dt = datetime(int(f"20{yy}"), int(mm), int(dd), tzinfo=timezone.utc)
        if as_of:
            ref = datetime.strptime(as_of, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        else:
            ref = datetime.now(timezone.utc)
        dte = max(0, (exp_dt - ref).days)
    except Exception:
        dte = 0
    return {"ticker": ticker, "option_type": option_type,
            "strike": strike, "expiry": expiry, "dte": dte}


def _fmt_prem(p: float) -> str:
    return f"${p/1_000_000:.1f}M" if p >= 1_000_000 else f"${p/1_000:.0f}K"


def _tkr_prem_str(snap: dict) -> str:
    """Compact 'C $1.2M (68%) / P $560K (32%)' for a ticker premium snapshot."""
    if not snap:
        return ""
    c, p = float(snap.get("call", 0)), float(snap.get("put", 0))
    tot = c + p
    if tot <= 0:
        return ""
    return (f"P {_fmt_prem(p)} ({p/tot*100:.0f}%) / "
            f"C {_fmt_prem(c)} ({c/tot*100:.0f}%)")


def _pct_str(v: float) -> str:
    return f"{v:+.0f}%"


def _median(vals: list) -> float:
    if not vals:
        return 0.0
    s = sorted(vals)
    n = len(s)
    return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2.0


def _pricing_warning() -> list:
    """Loud banner when pricing was unavailable for the whole run."""
    if _MASSIVE_AUTH_FAILED:
        return [
            "",
            "⚠️ OPTION PRICING UNAVAILABLE",
            "Every Massive key was rejected (401), so all pricing columns are",
            "blank. Check MASSIVE_API_KEYS / MASSIVE_API_KEY in Railway.",
        ]
    if _MASSIVE_BAD_KEYS:
        return [
            "",
            "⚠️ PARTIAL PRICING — one or more Massive keys are dead:",
            f"   rejected: {', '.join(_MASSIVE_BAD_KEYS)}",
            "Requests rotate across all keys, so roughly 1-in-N fetches hit the",
            "dead key and 401 — leaving scattered blank rows. Remove or replace",
            "that key in Railway and re-run for complete pricing.",
        ]
    return []


def _build_performance_summary(alerts: list) -> list:
    """
    Aggregate performance across every alert that has pricing data.

    Two win definitions, because they answer different questions:
      • Win @ expiry  — return at expiration > 0 (held to the end)
      • Win @ peak    — max gain reached TARGET% at any point before expiry
                        (closer to how you'd actually trade it)

    TARGETED_STRIKES_WIN_TARGET (default 50) sets the peak target %.

    Returns a list of text lines, or [] if no alert had pricing.
    """
    priced = [a for a in alerts if a.get("pricing")]
    if not priced:
        return []

    target = float(os.environ.get("TARGETED_STRIKES_WIN_TARGET", "50"))

    exp_rets   = [a["pricing"]["expiry_pct"]        for a in priced]
    max_gains  = [a["pricing"]["max_gain_pct"]      for a in priced]
    max_dds    = [a["pricing"]["max_dd_pct"]           for a in priced]
    givebacks  = [a["pricing"].get("giveback_pct", 0.0) for a in priced]
    life_dds   = [a["pricing"].get("mae_pct",
                                   a["pricing"]["max_draw_pct"]) for a in priced]

    n = len(priced)
    wins_exp    = sum(1 for r in exp_rets  if r > 0)
    wins_peak   = sum(1 for g in max_gains if g >= target)
    losses_exp  = n - wins_exp
    losses_peak = n - wins_peak

    avg = lambda xs: (sum(xs) / len(xs)) if xs else 0.0

    # Split average return by outcome — a 40% win rate with huge winners is a
    # very different system than 40% with small ones.
    win_rets  = [r for r in exp_rets if r > 0]
    lose_rets = [r for r in exp_rets if r <= 0]

    lines = [
        "",
        "━━━━━━ 📊 OVERALL PERFORMANCE ━━━━━━",
        f"Alerts with pricing: {n}",
        "",
        f"🎯 Win @ peak (≥{target:.0f}%): {wins_peak}/{n} "
        f"({100*wins_peak/n:.0f}%)  |  Lose: {losses_peak} ({100*losses_peak/n:.0f}%)",
        f"🏁 Win @ expiry:        {wins_exp}/{n} "
        f"({100*wins_exp/n:.0f}%)  |  Lose: {losses_exp} ({100*losses_exp/n:.0f}%)",
        "",
        f"📈 Avg max gain:        {_pct_str(avg(max_gains))}  "
        f"(median {_pct_str(_median(max_gains))})",
        f"📉 Avg max drawdown:    {_pct_str(avg(max_dds))}  "
        f"(median {_pct_str(_median(max_dds))})   ← below entry before peak",
        f"↩️  Avg give-back:       {_pct_str(avg(givebacks))}  "
        f"(median {_pct_str(_median(givebacks))})   ← from peak",
        f"💀 Avg life low vs entry: {_pct_str(avg(life_dds))}  "
        f"(median {_pct_str(_median(life_dds))})",
        f"🏁 Avg return @ expiry: {_pct_str(avg(exp_rets))}  "
        f"(median {_pct_str(_median(exp_rets))})",
    ]

    if win_rets:
        lines.append(f"   └ avg winner @ exp: {_pct_str(avg(win_rets))} "
                     f"(best {_pct_str(max(exp_rets))})")
    if lose_rets:
        lines.append(f"   └ avg loser  @ exp: {_pct_str(avg(lose_rets))} "
                     f"(worst {_pct_str(min(exp_rets))})")

    # Best / worst — useful for eyeballing outliers.
    best  = max(priced, key=lambda a: a["pricing"]["max_gain_pct"])
    worst = min(priced, key=lambda a: a["pricing"]["expiry_pct"])
    bo = "C" if best["direction"] == "call" else "P"
    wo = "C" if worst["direction"] == "call" else "P"
    lines += [
        "",
        f"🥇 Best runup:  ${best['ticker']} {best['strike']}{bo} "
        f"{_pct_str(best['pricing']['max_gain_pct'])}",
        f"🥉 Worst @ exp: ${worst['ticker']} {worst['strike']}{wo} "
        f"{_pct_str(worst['pricing']['expiry_pct'])}",
    ]

    # Split by direction — calls and puts often behave very differently.
    calls = [a for a in priced if a["direction"] == "call"]
    puts  = [a for a in priced if a["direction"] == "put"]
    if calls and puts:
        cw = sum(1 for a in calls if a["pricing"]["max_gain_pct"] >= target)
        pw = sum(1 for a in puts  if a["pricing"]["max_gain_pct"] >= target)
        lines += [
            "",
            f"📈 Calls: {len(calls)} alerts, {cw} hit +{target:.0f}% "
            f"({100*cw/len(calls):.0f}%), avg max "
            f"{_pct_str(avg([a['pricing']['max_gain_pct'] for a in calls]))}",
            f"📉 Puts:  {len(puts)} alerts, {pw} hit +{target:.0f}% "
            f"({100*pw/len(puts):.0f}%), avg max "
            f"{_pct_str(avg([a['pricing']['max_gain_pct'] for a in puts]))}",
        ]

    # Early-session split — does pre-cutoff flow perform differently?
    early_a = [a for a in priced if a.get("early")]
    late_a  = [a for a in priced if not a.get("early")]
    if early_a and late_a:
        ew = sum(1 for a in early_a if a["pricing"]["max_gain_pct"] >= target)
        lw = sum(1 for a in late_a  if a["pricing"]["max_gain_pct"] >= target)
        lines += [
            "",
            f"⏰ Early-session: {len(early_a)} alerts, {ew} hit +{target:.0f}% "
            f"({100*ew/len(early_a):.0f}%)",
            f"🕐 Regular:       {len(late_a)} alerts, {lw} hit +{target:.0f}% "
            f"({100*lw/len(late_a):.0f}%)",
        ]

    return lines


_MASSIVE_AUTH_FAILED = False   # True only when EVERY key is rejected
_MASSIVE_BAD_KEYS: list = []   # keys that returned 401/403, reported once


def _massive_key_health() -> tuple:
    """
    Probe EVERY configured Massive key and report which work.

    A single dead key in the rotation (e.g. a stale MASSIVE_API_KEY_3 or a
    legacy POLYGON_API_KEY) makes roughly 1-in-N requests 401 — producing
    scattered blank pricing that looks like missing data rather than an auth
    problem. Testing keys[0] alone would miss that entirely.

    Bypasses _massive_acquire: this is diagnosis, not data, and must not burn
    rate-limit slots. Returns (good_count, bad_labels).
    """
    good, bad = 0, []
    try:
        import bullflow_presets as _bp
        keys = _bp._massive_keys()
        if not keys:
            return 0, ["no keys configured"]
        # Label each key by which env var it most likely came from.
        labels = {}
        raw = os.environ.get("MASSIVE_API_KEYS", "").strip()
        if raw:
            for i, k in enumerate([x.strip() for x in raw.split(",") if x.strip()]):
                labels[k] = f"MASSIVE_API_KEYS[{i}]"
        for name in ("MASSIVE_API_KEY", "MASSIVE_API_KEY_2",
                     "MASSIVE_API_KEY_3", "POLYGON_API_KEY"):
            v = os.environ.get(name, "").strip()
            if v:
                labels.setdefault(v, name)

        base = os.environ.get("MASSIVE_BASE_URL", "https://api.massive.com")
        for k in keys:
            label = labels.get(k, f"...{k[-4:]}" if len(k) > 4 else "key")
            try:
                r = requests.get(f"{base}/v2/aggs/ticker/AAPL/prev",
                                 params={"apiKey": k}, timeout=10)
                if r.status_code in (401, 403):
                    bad.append(label)
                    print(f"[TARGETED_BT] Massive key {label}: "
                          f"REJECTED (HTTP {r.status_code})")
                else:
                    good += 1
                    print(f"[TARGETED_BT] Massive key {label}: OK "
                          f"(HTTP {r.status_code})")
            except Exception as e:
                print(f"[TARGETED_BT] Massive key {label}: probe error {e}")
    except Exception as e:
        print(f"[TARGETED_BT] key health check error: {e}")
    return good, bad


def _enrich_alert_pricing(alert: dict) -> None:
    """
    Enrich one fired alert with option-price stats over the life of the trade,
    from Massive 30-minute bars: entry (price on the triggering/most-recent
    flow), min low, max high, price at expiration, and the % moves from entry.

    Mutates the alert dict in place, adding an "pricing" sub-dict. Silent on
    failure (leaves pricing=None) so a data gap never breaks the backtest.

    Reuses preset_backtest._fetch_option_intraday (Massive, key rotation +
    rate limiting already handled there).
    """
    alert["pricing"] = None
    alert["pricing_note"] = ""
    fills = alert.get("fills", [])
    if not fills:
        alert["pricing_note"] = "no fills"
        return
    trigger = fills[-1]                       # the most-recent flow (fired the alert)
    occ         = trigger.get("occ", "")
    expiry_iso  = trigger.get("expiry_iso", "")
    entry_price = float(trigger.get("price", 0) or 0)
    if not occ or not expiry_iso or entry_price <= 0:
        alert["pricing_note"] = ("no OCC symbol" if not occ else
                                 "no expiry" if not expiry_iso else "no entry price")
        return

    start_iso = alert.get("date", "")         # fire day (YYYY-MM-DD)
    if not start_iso:
        alert["pricing_note"] = "no alert date"
        return

    global _MASSIVE_AUTH_FAILED, _MASSIVE_BAD_KEYS
    if _MASSIVE_AUTH_FAILED:
        alert["pricing_note"] = "skipped — all Massive keys rejected (401)"
        return

    try:
        from preset_backtest import _fetch_option_intraday
        bars = _fetch_option_intraday(occ, start_iso, expiry_iso)
    except Exception as e:
        print(f"[TARGETED_BT] pricing fetch error {occ}: {e}")
        alert["pricing_note"] = f"fetch error: {type(e).__name__}"
        return

    # An empty result right after a fresh key check almost always means auth.
    # Probe once so we can fail the whole pass fast instead of sleeping 60s
    # between every doomed request.
    if not bars and not _MASSIVE_BAD_KEYS:
        good, bad = _massive_key_health()
        if bad:
            _MASSIVE_BAD_KEYS = bad
        if good == 0:
            _MASSIVE_AUTH_FAILED = True
            alert["pricing_note"] = "all Massive keys rejected (401)"
            return
    if not bars:
        alert["pricing_note"] = "no Massive bars for contract"
        return

    # Bullflow sends epoch in seconds OR milliseconds; Massive bar ts is in
    # seconds. Normalize before comparing, or the filter silently drops every
    # bar and falls back to the full series (including pre-alert bars).
    try:
        from bullflow_presets import _norm_epoch
        fire_ts = _norm_epoch(trigger.get("ts", 0))
    except Exception:
        fire_ts = float(trigger.get("ts", 0) or 0)
        if fire_ts > 1e11:
            fire_ts /= 1000.0

    after = [b for b in bars if b.get("ts", 0) >= fire_ts]
    if after:
        life = after
    else:
        # No bars at/after the alert (e.g. fired in the last 30 min of expiry
        # day). Fall back to the full series but flag it — the stats then
        # cover the contract's whole day, not just post-alert.
        life = bars
        alert["pricing_note"] = "no post-alert bars; stats cover full window"

    highs = [b["high"] for b in life if b.get("high")]
    lows  = [b["low"]  for b in life if b.get("low")]
    if not highs or not lows:
        alert["pricing_note"] = "bars had no high/low"
        return

    max_px = max(highs)
    min_px = min(lows)
    exp_px = life[-1]["close"]                # close of the last bar (expiry day)

    # MAX DRAWDOWN: how far below ENTRY it traded BEFORE reaching its peak.
    # This is the heat you'd have to sit through while still holding for the
    # run — the number that decides whether you'd have survived the trade.
    peak_idx = max(range(len(life)),
                   key=lambda i: life[i].get("high", 0) or 0)
    pre_peak_lows = [b["low"] for b in life[:peak_idx + 1] if b.get("low")]
    dd_low = min(pre_peak_lows) if pre_peak_lows else min_px

    # Give-back after the peak (peak -> lowest low that follows it). Useful
    # for "could I have held the winner", but NOT the drawdown.
    post_peak_lows = [b["low"] for b in life[peak_idx:] if b.get("low")]
    giveback_low = min(post_peak_lows) if post_peak_lows else max_px

    def pct(a, b):
        return ((a - b) / b * 100.0) if b else 0.0

    alert["pricing"] = {
        "entry":       entry_price,
        "min":         min_px,
        "max":         max_px,
        "at_expiry":   exp_px,
        # THE drawdown: lowest point below entry before the peak.
        "dd_low":      dd_low,
        "max_dd_pct":  pct(dd_low, entry_price),
        # Back-compat aliases — same value, older keys.
        "pre_peak_min":      dd_low,
        "pre_peak_draw_pct": pct(dd_low, entry_price),
        "max_gain_pct":  pct(max_px, entry_price),
        # Post-peak give-back (peak -> subsequent low). Not a drawdown.
        "giveback_low":  giveback_low,
        "giveback_pct":  pct(giveback_low, max_px),
        # Worst level vs entry across the whole life (incl. expiry decay).
        "mae_pct":       pct(min_px, entry_price),
        "max_draw_pct":  pct(min_px, entry_price),
        "expiry_pct":    pct(exp_px, entry_price),
        "bars":        len(life),
    }


def _stream_one_day(date: str, api_key: str, max_attempts: int = 4) -> tuple[list, int, int, dict]:
    """
    Streams a single day and returns (alerts_fired, event_count,
    targeted_events, seen_names). Isolated state — never touches
    production _TARGETED.

    Bullflow's backtest SSE endpoint replays a finite day and does NOT
    support resume-from-offset, so if the connection drops mid-stream we
    restart the whole day from scratch (state reset each attempt) rather
    than reconnecting — this avoids double-counting. Retries up to
    max_attempts times on premature drops / transient network errors.
    """
    url = _build_url(api_key, date)
    last_err = None

    for attempt in range(1, max_attempts + 1):
        try:
            return _stream_one_day_once(url, date)
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout,
                requests.exceptions.HTTPError) as e:
            last_err = e
            if attempt < max_attempts:
                wait = 15 * attempt   # 15s, 30s, 45s backoff
                print(f"[TARGETED_BT] Stream dropped on attempt {attempt}/{max_attempts} "
                      f"({type(e).__name__}: {e}) — restarting day in {wait}s")
                time.sleep(wait)
            else:
                print(f"[TARGETED_BT] Stream failed after {max_attempts} attempts: {e}")
        except Exception as e:
            # Non-network error — don't retry blindly, surface it.
            raise
    # Exhausted retries — re-raise the last network error for the caller to report.
    raise last_err


def _stream_one_day_once(url: str, date: str) -> tuple[list, int, int, dict]:
    """One streaming attempt. Fresh state; raises on premature drop."""
    bt_state:        dict = {}   # ticker_direction → {"strike","expiry","fills":[...],"last_alerted_count"}
    alerts_fired:    list = []
    event_count      = 0
    targeted_events  = 0
    seen_names:      dict = {}
    tkr_prem:        dict = {}   # ticker → running call/put premium (whole day)
    tkr_prem_330:    dict = {}   # ticker → call/put premium as of 15:30 ET

    resp = requests.get(url, stream=True, timeout=660,
                        headers={"Accept": "text/event-stream",
                                 "Cache-Control": "no-cache"})
    resp.raise_for_status()

    buffer = ""
    for chunk in resp.iter_content(chunk_size=2048, decode_unicode=True):
        if not chunk:
            continue
        buffer += chunk

        while "\n\n" in buffer:
            raw_event, buffer = buffer.split("\n\n", 1)
            data_lines = [l[6:] for l in raw_event.split("\n") if l.startswith("data: ")]
            if not data_lines:
                continue
            try:
                data = json.loads(data_lines[0])
            except Exception:
                continue

            event_count += 1
            inner      = data.get("data", data)
            alert_name = inner.get("alertName", "")
            seen_names[alert_name] = seen_names.get(alert_name, 0) + 1

            name_match = (alert_name == TARGETED_FILTER or
                          alert_name.lower() == TARGETED_FILTER.lower())
            if not name_match:
                continue
            targeted_events += 1

            symbol = inner.get("symbol", inner.get("ticker", ""))
            parsed = _parse_occ(symbol, as_of=date) if "O:" in str(symbol) else None
            if not parsed:
                raw_ticker = (inner.get("ticker", "") or "").upper()[:10]
                if not raw_ticker:
                    continue
                parsed = {
                    "ticker":      raw_ticker,
                    "option_type": inner.get("optionType", "call"),
                    "strike":      str(inner.get("strikePrice", "?")),
                    "expiry":      inner.get("expirationDate", "?"),
                    "dte":         0,
                }

            direction = "call" if "call" in str(parsed["option_type"]).lower() else "put"
            ticker    = parsed["ticker"]
            strike    = parsed["strike"]
            expiry    = parsed["expiry"]
            price     = float(inner.get("averageFillPrice") or inner.get("tradePrice") or 0)
            premium   = float(inner.get("alertPremium") or 0)
            is_sweep  = str(inner.get("alertFillType", "")).upper() in ("FULL_ASK", "AA")
            stock_px  = float(inner.get("stockPrice") or 0)

            # Normalized OCC symbol (e.g. GOOGL260717C00370000) + ISO expiry,
            # used later to pull 30M option bars from Massive for pricing.
            occ_norm = ""
            expiry_iso = ""
            _mm = re.search(r'O:([A-Z]+\d{6}[CP]\d+)', str(symbol))
            if _mm:
                occ_norm = _mm.group(1)
                _e = re.search(r'[A-Z]+(\d{2})(\d{2})(\d{2})[CP]', occ_norm)
                if _e:
                    expiry_iso = f"20{_e.group(1)}-{_e.group(2)}-{_e.group(3)}"

            event_ts = float(inner.get("timestamp") or time.time())
            est_str  = str(inner.get("estTimestamp", ""))
            time_str = est_str[11:19] if len(est_str) >= 19 else datetime.now(ET).strftime("%-I:%M:%S %p")
            early    = _is_early_str(est_str)

            # Per-ticker call/put premium — EVERY Targeted_Strikes_Expiry fill
            # for this ticker (all strikes/expiries the filter surfaces).
            # Deliberately placed BEFORE the SKIP_EARLY drop: the cutoff exists
            # to shape run detection, not to shrink this context measure, and
            # the live path tallies every fill too. Also snapshot as of 15:30 ET.
            if premium > 0:
                _row = tkr_prem.setdefault(ticker, {"call": 0.0, "put": 0.0,
                                                     "call_n": 0, "put_n": 0})
                _row[direction] += premium
                _row[f"{direction}_n"] += 1
                if len(est_str) >= 16 and est_str[11:16] < "15:30":
                    _r330 = tkr_prem_330.setdefault(ticker, {"call": 0.0, "put": 0.0,
                                                              "call_n": 0, "put_n": 0})
                    _r330[direction] += premium
                    _r330[f"{direction}_n"] += 1

            if SKIP_EARLY and early:
                continue   # drop pre-cutoff fill entirely — not counted, not stored

            key = f"{ticker}_{direction}"   # streak is per ticker+direction
            fill = {
                "strike": strike, "expiry": expiry, "price": price,
                "premium": premium, "sweep": is_sweep, "dte": parsed["dte"],
                "stock_px": stock_px, "time": time_str, "ts": event_ts, "early": early,
                "occ": occ_norm, "expiry_iso": expiry_iso,
            }

            streak = bt_state.get(key)
            same_contract = (streak is not None
                             and streak.get("strike") == strike
                             and streak.get("expiry") == expiry)

            if same_contract:
                # Continue the current consecutive run.
                streak["fills"].append(fill)
            else:
                # First fill for this ticker+dir, OR a same-direction fill at a
                # DIFFERENT strike/expiry -> previous run breaks; start fresh.
                # (Opposite direction / other tickers live under their own key,
                #  so they never reach here and never break this streak.)
                streak = {"strike": strike, "expiry": expiry,
                          "fills": [fill], "last_alerted_count": 0}
                bt_state[key] = streak

            fills        = streak["fills"]
            count        = len(fills)
            last_alerted = streak["last_alerted_count"]

            # GATE_UNTIL_CUTOFF: hold the alert while the triggering fill is
            # pre-cutoff (early=True); don't advance last_alerted so a later
            # at/after-cutoff fill fires with the full accumulated count.
            gated = GATE_UNTIL_CUTOFF and early

            if count >= THRESHOLD and count > last_alerted and not gated:
                streak["last_alerted_count"] = count
                _snap = dict(tkr_prem.get(ticker,
                             {"call": 0.0, "put": 0.0, "call_n": 0, "put_n": 0}))
                alerts_fired.append({
                    "ticker":     ticker,
                    "strike":     strike,
                    "expiry":     expiry,
                    "direction":  direction,
                    "fills":      list(fills),
                    "count":      count,
                    "total_prem": sum(f["premium"] for f in fills),
                    "is_addon":   last_alerted > 0,
                    "early":      any(f.get("early") for f in fills),
                    "date":       date,
                    "time":       time_str,
                    "tkr_at_alert": _snap,
                })

    # Attach the ticker's 15:30 ET call/put premium to every alert, so each
    # alert can be read as "how it looked at fire time" vs "how the day stood
    # at 3:30". Falls back to the whole-day tally if no pre-15:30 flow.
    for a in alerts_fired:
        a["tkr_at_330"] = dict(tkr_prem_330.get(
            a["ticker"], {"call": 0.0, "put": 0.0, "call_n": 0, "put_n": 0}))

    return alerts_fired, event_count, targeted_events, seen_names


def _run_backtest_thread(date: str, bot_token: str, chat_id: str, detail: bool = False):
    from sms import send_telegram

    api_key = os.environ.get("BULLFLOW_API_KEY", "")
    if not api_key:
        send_telegram("❌ BULLFLOW_API_KEY not set", bot_token, chat_id)
        return

    try:
        alerts_fired, event_count, targeted_events, seen_names = _stream_one_day(date, api_key)
    except (requests.exceptions.ChunkedEncodingError,
            requests.exceptions.ConnectionError,
            requests.exceptions.Timeout) as e:
        send_telegram(
            f"❌ Backtest for {date} failed: Bullflow stream kept dropping "
            f"after several retries ({type(e).__name__}). This is usually a "
            f"transient Bullflow/network issue — try again in a minute.",
            bot_token, chat_id)
        return
    except Exception as e:
        send_telegram(f"❌ Backtest error ({date}): {e}", bot_token, chat_id)
        return

    _report_results(date, alerts_fired, event_count, targeted_events, seen_names,
                     bot_token, chat_id, detail)


def _report_results(date: str, alerts_fired: list, event_count: int, targeted_events: int,
                     seen_names: dict, bot_token: str, chat_id: str, detail: bool):
    from sms import send_telegram

    if not alerts_fired:
        top_names = sorted(seen_names.items(), key=lambda x: -x[1])[:8]
        names_str = "\n".join(f"  {n!r}: {c}" for n, c in top_names) if top_names else "  (none)"
        send_telegram(
            f"🎯 Targeted Strikes Backtest: {date}\n"
            f"No alerts found for filter: {TARGETED_FILTER!r}\n"
            f"({targeted_events} matching events of {event_count} total)\n\n"
            f"Alert names seen in stream:\n{names_str}\n\n"
            f"If your filter name differs, set TARGETED_STRIKES_FILTER_NAME in Railway.",
            bot_token, chat_id)
        return

    # One line per unique key that crossed threshold — LATEST/highest count per key
    latest_per_key = {}
    for a in alerts_fired:
        latest_per_key[f"{a['ticker']}_{a['strike']}_{a['expiry']}_{a['direction']}"] = a
    early_count = sum(1 for a in alerts_fired if a["early"])

    # Enrich each displayed strike with option pricing from Massive 30M bars.
    if os.environ.get("TARGETED_STRIKES_PRICING", "true").lower() in ("true","1","yes","on"):
        for a in latest_per_key.values():
            _enrich_alert_pricing(a)

    # Point-in-time swing scoring (no lookahead — see module docstring).
    if os.environ.get("TARGETED_SWING_BACKTEST_SCORE", "true").lower() in ("true","1","yes","on"):
        try:
            from targeted_swing_backtest_score import score_historical_alert
            for a in latest_per_key.values():
                score_historical_alert(a)
        except Exception as _se:
            print(f"[SWINGBT] scoring error: {_se}")

    summary = [
        f"🎯 Targeted Strikes Backtest: {date}",
        f"━━━ {len(latest_per_key)} strikes crossed {THRESHOLD}x "
        f"| {len(alerts_fired)} total alerts ({early_count} early-session) "
        f"| {targeted_events} events ━━━",
        "",
    ]
    for a in sorted(latest_per_key.values(), key=lambda x: -x["count"]):
        emoji  = "📈" if a["direction"] == "call" else "📉"
        otype  = "C" if a["direction"] == "call" else "P"
        early_s = " ⏰EARLY" if a["early"] else ""
        fills   = a.get("fills", [])
        # time the run first crossed THRESHOLD = the THRESHOLD-th fill's time
        cross_time = (fills[THRESHOLD - 1].get("time") or a["time"]) if len(fills) >= THRESHOLD else a["time"]
        latest_time = a["time"]
        if a["count"] > THRESHOLD:
            time_str = f"crossed {THRESHOLD}x @ {cross_time}, latest @ {latest_time}"
        else:
            time_str = f"@ {cross_time}"
        summary.append(
            f"{emoji} ${a['ticker']} {a['strike']}{otype} {a['expiry']}  "
            f"{a['count']}x  {_fmt_prem(a['total_prem'])}  {time_str}{early_s}"
        )
        if a.get("swing_score"):
            summary.append(f"   🎯 swing score {a['swing_score']:.0f}/100"
                           + (f" — {a['swing_notes'][0]}" if a.get("swing_notes") else ""))
        elif a.get("swing_dq"):
            summary.append(f"   🎯 not swing-eligible: {a['swing_dq']}")
        pr = a.get("pricing")
        if pr:
            summary.append(
                f"   💵 entry ${pr['entry']:.2f} → max ${pr['max']:.2f} "
                f"({pr['max_gain_pct']:+.0f}%)"
            )
            summary.append(
                f"      max DD {pr['max_dd_pct']:+.0f}% "
                f"(low ${pr['dd_low']:.2f} before peak) "
                f"| exp ${pr['at_expiry']:.2f} ({pr['expiry_pct']:+.0f}%)"
            )
            summary.append(
                f"      give-back after peak {pr['giveback_pct']:+.0f}% "
                f"(${pr['max']:.2f}→${pr['giveback_low']:.2f}) "
                f"| life low ${pr['min']:.2f} ({pr['mae_pct']:+.0f}%)"
            )
        _ta = _tkr_prem_str(a.get("tkr_at_alert"))
        _t3 = _tkr_prem_str(a.get("tkr_at_330"))
        if _ta:
            summary.append(f"   📊 {a['ticker']} @alert: {_ta}")
        if _t3:
            summary.append(f"   📊 {a['ticker']} @3:30:  {_t3}")
    summary += _pricing_warning()
    summary += _build_performance_summary(list(latest_per_key.values()))
    try:
        from targeted_swing_backtest_score import score_correlation_summary
        summary += score_correlation_summary(list(latest_per_key.values()))
    except Exception:
        pass
    send_telegram("\n".join(summary), bot_token, chat_id)

    if detail:
        from targeted_strikes_tracker import build_targeted_strikes_alert
        for a in latest_per_key.values():
            result = {
                "ticker": a["ticker"], "strike": a["strike"], "expiry": a["expiry"],
                "direction": a["direction"], "fills": a["fills"], "count": a["count"],
                "total_prem": a["total_prem"], "is_addon": a["is_addon"],
                "early": a["early"], "threshold": THRESHOLD,
            }
            send_telegram(build_targeted_strikes_alert(result), bot_token, chat_id)


def _build_range_workbook(all_alerts: list, start_date: str, end_date: str) -> str:
    """One row per alert (including add-ons), sorted by date/time. Returns .xlsx path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Targeted Strikes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    center      = Alignment(horizontal="center")

    from openpyxl.utils import get_column_letter

    cols = ["Date", f"Crossed {THRESHOLD}x Time", "Fill Time", "Ticker", "Direction",
            "Strike", "Expiry", "Count", "Add-On", "Early Session", "Combined Premium",
            "Entry", "Min", "Max", "At Expiry", "Max Gain %", "Max DD %", "DD Low", "Give-Back %", "Give-Back Low", "Life Low % (vs entry)", "Expiry %",
            "Pricing Note",
            "Tkr Put Prem @Alert", "Tkr Call Prem @Alert", "Tkr Put% @Alert",
            "Tkr Put Prem @3:30", "Tkr Call Prem @3:30", "Tkr Put% @3:30",
            "Swing Score", "Swing DQ", "Score: 30M", "Score: Daily", "Score: Flow"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center

    for a in sorted(all_alerts, key=lambda x: (x["date"], x["time"])):
        fills = a.get("fills", [])
        cross_time = (fills[THRESHOLD - 1].get("time") or a["time"]) if len(fills) >= THRESHOLD else a["time"]
        pr = a.get("pricing") or {}
        _ta = a.get("tkr_at_alert") or {}
        _t3 = a.get("tkr_at_330") or {}
        _sb = a.get("swing_breakdown") or {}
        ws.append([
            a["date"], cross_time, a["time"], a["ticker"], a["direction"].upper(),
            a["strike"], a["expiry"], a["count"],
            "YES" if a["is_addon"] else "",
            "YES" if a["early"] else "",
            round(a["total_prem"], 2),
            round(pr["entry"], 2)      if pr else "",
            round(pr["min"], 2)        if pr else "",
            round(pr["max"], 2)        if pr else "",
            round(pr["at_expiry"], 2)  if pr else "",
            round(pr["max_gain_pct"], 1) if pr else "",
            round(pr.get("max_dd_pct", 0), 1)   if pr else "",
            round(pr.get("dd_low", 0), 2)       if pr else "",
            round(pr.get("giveback_pct", 0), 1) if pr else "",
            round(pr.get("giveback_low", 0), 2) if pr else "",
            round(pr.get("mae_pct", pr.get("max_draw_pct", 0)), 1) if pr else "",
            round(pr["expiry_pct"], 1)   if pr else "",
            (a.get("pricing_note") or "")
            if pr or a.get("pricing_note") is not None
            else "not priced (add-on row)",
            round(_ta.get("put", 0), 2), round(_ta.get("call", 0), 2),
            round(_ta["put"] / (_ta["call"] + _ta["put"]) * 100, 1)
            if (_ta.get("call", 0) + _ta.get("put", 0)) else "",
            round(_t3.get("put", 0), 2), round(_t3.get("call", 0), 2),
            round(_t3["put"] / (_t3["call"] + _t3["put"]) * 100, 1)
            if (_t3.get("call", 0) + _t3.get("put", 0)) else "",
            a.get("swing_score") or "",
            a.get("swing_dq") or "",
            _sb.get("m30", ""), _sb.get("daily", ""), _sb.get("flow", ""),
        ])

    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

    path = f"/tmp/targeted_strikes_{start_date}_to_{end_date}.xlsx"
    wb.save(path)
    return path


def _run_range_thread(start_date: str, end_date: str, bot_token: str, chat_id: str, detail: bool = False):
    from sms import send_telegram

    api_key = os.environ.get("BULLFLOW_API_KEY", "")
    if not api_key:
        send_telegram("❌ BULLFLOW_API_KEY not set", bot_token, chat_id)
        return

    start = datetime.strptime(start_date, "%Y-%m-%d")
    end   = datetime.strptime(end_date, "%Y-%m-%d")
    if end < start:
        send_telegram("❌ End date is before start date", bot_token, chat_id)
        return
    span = (end - start).days
    if span > MAX_RANGE_DAYS:
        send_telegram(f"❌ Range too large ({span} days). Max is {MAX_RANGE_DAYS} days (a full month).",
                       bot_token, chat_id)
        return

    all_alerts, total_events, total_targeted = [], 0, 0
    day = start
    days_run = 0
    while day <= end:
        if day.weekday() < 5:  # skip weekends
            date_str = day.strftime("%Y-%m-%d")
            try:
                alerts, events, targeted, _ = _stream_one_day(date_str, api_key)
                all_alerts.extend(alerts)
                total_events += events
                total_targeted += targeted
                days_run += 1
            except Exception as e:
                send_telegram(f"⚠️ Skipped {date_str}: {e}", bot_token, chat_id)
        day += timedelta(days=1)

    range_label = f"{start_date} → {end_date} ({days_run} trading days)"
    if not all_alerts:
        send_telegram(
            f"🎯 Targeted Strikes Range Backtest: {range_label}\n"
            f"No alerts found for filter {TARGETED_FILTER!r} "
            f"({total_targeted} matching events of {total_events} total).",
            bot_token, chat_id)
        return

    latest_per_key = {}
    for a in all_alerts:
        latest_per_key[f"{a['date']}_{a['ticker']}_{a['strike']}_{a['expiry']}_{a['direction']}"] = a
    early_count = sum(1 for a in all_alerts if a["early"])

    # Enrich the deduped strikes with Massive pricing. These are the same dict
    # objects that live in all_alerts, so the workbook picks up pricing too.
    if os.environ.get("TARGETED_STRIKES_PRICING", "true").lower() in ("true","1","yes","on"):
        for a in latest_per_key.values():
            _enrich_alert_pricing(a)

    # Point-in-time swing scoring (no lookahead — see module docstring).
    if os.environ.get("TARGETED_SWING_BACKTEST_SCORE", "true").lower() in ("true","1","yes","on"):
        try:
            from targeted_swing_backtest_score import score_historical_alert
            for a in latest_per_key.values():
                score_historical_alert(a)
        except Exception as _se:
            print(f"[SWINGBT] scoring error: {_se}")
    by_ticker: dict = {}
    for a in latest_per_key.values():
        by_ticker[a["ticker"]] = by_ticker.get(a["ticker"], 0) + 1

    top_tickers = sorted(by_ticker.items(), key=lambda x: -x[1])[:10]
    summary = [
        f"🎯 Targeted Strikes Range Backtest: {range_label}",
        f"━━━ {len(latest_per_key)} distinct strikes crossed {THRESHOLD}x "
        f"| {len(all_alerts)} total alerts ({early_count} early-session) ━━━",
        "",
        "Top tickers: " + ", ".join(f"{t}({c})" for t, c in top_tickers),
        "",
    ]
    for a in sorted(latest_per_key.values(), key=lambda x: (x["date"], -x["count"]))[:40]:
        emoji  = "📈" if a["direction"] == "call" else "📉"
        otype  = "C" if a["direction"] == "call" else "P"
        early_s = " ⏰" if a["early"] else ""
        fills   = a.get("fills", [])
        cross_time = (fills[THRESHOLD - 1].get("time") or a["time"]) if len(fills) >= THRESHOLD else a["time"]
        if a["count"] > THRESHOLD:
            time_str = f"{THRESHOLD}x@{cross_time} → {a['count']}x@{a['time']}"
        else:
            time_str = f"@{cross_time}"
        sc = ""
        if a.get("swing_score"):
            sc = f"  🎯{a['swing_score']:.0f}"
        elif a.get("swing_dq"):
            sc = "  🎯dq"
        summary.append(
            f"{a['date']} {emoji} ${a['ticker']} {a['strike']}{otype} {a['expiry']}  "
            f"{a['count']}x  {_fmt_prem(a['total_prem'])}  {time_str}{early_s}{sc}"
        )
    if len(latest_per_key) > 40:
        summary.append(f"... and {len(latest_per_key) - 40} more — see attached workbook")

    summary += _pricing_warning()
    summary += _build_performance_summary(list(latest_per_key.values()))
    try:
        from targeted_swing_backtest_score import score_correlation_summary
        summary += score_correlation_summary(list(latest_per_key.values()))
    except Exception:
        pass

    send_telegram("\n".join(summary), bot_token, chat_id)

    # Single-factor A/B analysis — only meaningful over a wide range.
    try:
        from factor_lab import run_factor_lab, save_to_pool, load_pool, pool_summary
        this_run = list(latest_per_key.values())
        # Bullflow caps a replay at 31 days, so accumulate across runs — a
        # single month is far too small for the factor tests to have power.
        n_pool = save_to_pool(this_run)
        pooled = load_pool()
        msg = []
        if n_pool:
            msg += pool_summary() + [""]
        # Analyse the POOL when it is bigger than this run alone.
        target = pooled if len(pooled) > len(this_run) else this_run
        if len(pooled) > len(this_run):
            msg += [f"(analysing pooled {len(pooled)} alerts, not just this month)", ""]
        msg += run_factor_lab(target)
        send_telegram("\n".join(msg), bot_token, chat_id)
    except Exception as _fe:
        print(f"[FACTORLAB] error: {_fe}")

    try:
        from sms import send_telegram_document
        xlsx_path = _build_range_workbook(all_alerts, start_date, end_date)
        send_telegram_document(
            xlsx_path, bot_token, chat_id,
            caption=f"Targeted Strikes: {range_label} — every alert, one row each"
        )
    except Exception as e:
        send_telegram(f"⚠️ Could not build/send workbook: {e}", bot_token, chat_id)


def start_backtest(date: str, bot_token: str, chat_id: str, detail: bool = False) -> bool:
    """Validate date format and launch a single-day backtest in a background thread."""
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date):
        return False
    t = threading.Thread(
        target=_run_backtest_thread,
        args=(date, bot_token, chat_id, detail),
        daemon=True,
        name=f"targeted_strikes_backtest_{date}",
    )
    t.start()
    return True


_QUEUE_RUNNING = False


def _month_bounds(ym: str) -> tuple:
    """'2026-01' -> ('2026-01-01', '2026-01-31'), clamped to today."""
    y, m = int(ym[:4]), int(ym[5:7])
    start = datetime(y, m, 1)
    end = (datetime(y + (m == 12), (m % 12) + 1, 1) - timedelta(days=1))
    today = datetime.now()
    if end > today:
        end = today
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def _run_months_thread(months: list, bot_token: str, chat_id: str, detail: bool):
    """
    Run each month to completion, one at a time, in ONE worker thread.

    Sequential on purpose: concurrent ranges share the same Massive budget
    (2 keys x 5/min) and the 10/min dark pool limit, so running them in
    parallel is slower overall AND makes rate-limit failures more likely.
    Each month saves to the factor pool as it finishes, so a crash or a
    container restart only costs the month in flight — not the whole batch.
    """
    global _QUEUE_RUNNING
    from sms import send_telegram
    _QUEUE_RUNNING = True
    done, failed = [], []
    try:
        send_telegram(
            f"📋 Queued {len(months)} month(s): {', '.join(months)}\n"
            f"Running one at a time — each saves to the factor pool as it "
            f"finishes, so progress survives a restart.\n"
            f"Expect roughly 1.5-2.5h per month.",
            bot_token, chat_id)
        for i, ym in enumerate(months, 1):
            s, e = _month_bounds(ym)
            print(f"[QUEUE] ({i}/{len(months)}) {ym}: {s} → {e}")
            try:
                _run_range_thread(s, e, bot_token, chat_id, detail)
                done.append(ym)
            except Exception as ex:
                print(f"[QUEUE] {ym} failed: {ex}")
                failed.append(ym)
                send_telegram(f"⚠️ {ym} failed: {ex}\nContinuing with the rest.",
                              bot_token, chat_id)
        lines = [f"✅ Queue finished — {len(done)}/{len(months)} month(s) done"]
        if failed:
            lines.append(f"❌ Failed: {', '.join(failed)} (re-run these; "
                         f"duplicates are deduped automatically)")
        try:
            from factor_lab import pool_summary
            lines += [""] + pool_summary()
            lines += ["", "Run /factor_lab to analyse everything pooled."]
        except Exception:
            pass
        send_telegram("\n".join(lines), bot_token, chat_id)
    finally:
        _QUEUE_RUNNING = False


def start_backtest_months(months: list, bot_token: str, chat_id: str,
                          detail: bool = False) -> str:
    """
    Queue whole months to run back-to-back. `months` are 'YYYY-MM' strings.
    Returns "" on success or an error message.
    """
    if _QUEUE_RUNNING:
        return "A queue is already running — wait for it to finish."
    clean = []
    for m in months:
        m = str(m).strip()
        if not re.match(r'^\d{4}-\d{2}$', m):
            return f"Bad month format: '{m}' — use YYYY-MM (e.g. 2026-01)"
        clean.append(m)
    if not clean:
        return "No months given."
    if len(clean) > 12:
        return "Max 12 months per queue."
    threading.Thread(target=_run_months_thread,
                     args=(clean, bot_token, chat_id, detail),
                     daemon=True, name="targeted_month_queue").start()
    return ""


def start_backtest_range(start_date: str, end_date: str, bot_token: str, chat_id: str, detail: bool = False) -> bool:
    """Validate date formats and launch a multi-day backtest in a background thread."""
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', start_date) or not re.match(r'^\d{4}-\d{2}-\d{2}$', end_date):
        return False
    t = threading.Thread(
        target=_run_range_thread,
        args=(start_date, end_date, bot_token, chat_id, detail),
        daemon=True,
        name=f"targeted_strikes_range_{start_date}_{end_date}",
    )
    t.start()
    return True

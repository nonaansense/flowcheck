"""
preset_sweep.py — Parameter sweep for TP1 / TP2 / trailing-stop.

Streams a date range ONCE, collects every qualifying preset alert and its
option price path, then re-runs the 2-contract leg simulation across a grid
of TP1 / TP2 / trail-offset combinations — WITHOUT re-fetching any data
(option bars are cached by preset_backtest). Results go to an Excel ranked
by total P/L so you can see which parameters actually had an edge.

The point: your live settings are one cell in that grid. The sweep shows you
whether the neighbours are better, and by how much.

Telegram: /preset_sweep YYYY-MM-DD YYYY-MM-DD

Grid (override with env vars, comma-separated):
  SWEEP_TP1_PCTS   = 0.50,0.75,1.01,1.25,1.50
  SWEEP_TP2_PCTS   = 1.00,1.50,2.01,2.50
  SWEEP_TRAIL_PCTS = 0.30,0.40,0.50,0.60,0.75
"""
import os, re, threading
from datetime import datetime

MAX_RANGE_DAYS = 31


def _pcts(env_name: str, default: str) -> list:
    raw = os.environ.get(env_name, default)
    out = []
    for p in raw.split(","):
        p = p.strip()
        if not p:
            continue
        try:
            out.append(float(p))
        except Exception:
            continue
    return out


def _grid() -> tuple:
    entries = _pcts("SWEEP_ENTRY_PCTS", "0.00,0.10,0.15,0.20,0.25,0.30,0.40")
    tp1s    = _pcts("SWEEP_TP1_PCTS",   "0.50,0.75,1.01,1.25,1.50")
    tp2s    = _pcts("SWEEP_TP2_PCTS",   "1.00,1.50,2.01,2.50")
    trails  = _pcts("SWEEP_TRAIL_PCTS", "0.30,0.40,0.50,0.60,0.75")
    return entries, tp1s, tp2s, trails


def _score_combo(trades: list, entry_pct: float, tp1_pct: float, tp2_pct: float,
                 trail_pct: float) -> dict:
    """
    Re-simulate every ALERT with one parameter combination.

    Crucially, the entry discount is applied HERE — so a deeper discount
    produces a better entry price but fewer fills. `trades` carry the FULL
    option path from the alert onward, not a pre-filtered post-fill window.

    A never-filled alert contributes $0 and is counted as a miss. The metric
    that makes discounts comparable is **P/L per ALERT** (total ÷ all alerts),
    which prices in the trades you didn't get.
    """
    from preset_backtest import _run_legs
    import bullflow_presets as bp

    total_usd = 0.0
    pcts, wins, t1_hits, t2_hits, fills = [], 0, 0, 0, 0
    by_day: dict = {}
    n_alerts = len(trades)

    for t in trades:
        flow = t["flow_price"]
        path = t["path"]
        if flow <= 0 or not path:
            continue

        entry = bp._round_up_tenth(flow * (1 - entry_pct))
        if entry <= 0:
            continue

        # Where would THIS entry have filled?
        fill_idx = None
        for i, b in enumerate(path):
            if b["low"] > 0 and b["low"] <= entry:
                fill_idx = i
                break
        if fill_idx is None:
            continue          # never filled at this discount — a miss

        fills += 1
        window = path[fill_idx:]
        t1 = bp._floor_cent(entry * (1 + tp1_pct))
        t2 = bp._floor_cent(entry * (1 + tp2_pct))
        offset = round(flow * trail_pct, 2)
        if offset <= 0:
            continue

        res = _run_legs(window, entry, offset, t1, t2, use_trail=True)
        total_usd += res["pnl_usd"]
        pcts.append(res["pnl_pct"])
        by_day[t["date"]] = round(by_day.get(t["date"], 0.0) + res["pnl_usd"], 2)
        if res["pnl_usd"] > 0:
            wins += 1
        if res["leg1_reason"] == "TP1":
            t1_hits += 1
        if res["leg2_reason"] == "TP2":
            t2_hits += 1

    n = len(pcts)
    day_vals   = list(by_day.values())
    green_days = sum(1 for v in day_vals if v > 0)
    best_day   = max(day_vals) if day_vals else 0.0
    concentration = round(best_day / total_usd * 100, 1) if total_usd > 0 else 0.0

    return {
        "entry_pct": round(entry_pct * 100, 0),
        "tp1_pct":   round(tp1_pct * 100, 0),
        "tp2_pct":   round(tp2_pct * 100, 0),
        "trail_pct": round(trail_pct * 100, 0),
        "alerts":    n_alerts,
        "fills":     fills,
        "fill_rate": round(fills / n_alerts * 100, 1) if n_alerts else 0.0,
        "trades":    n,
        "total_usd": round(total_usd, 2),
        # THE comparable metric across discounts — prices in the missed trades
        "usd_per_alert": round(total_usd / n_alerts, 2) if n_alerts else 0.0,
        "usd_per_fill":  round(total_usd / n, 2) if n else 0.0,
        "avg_pct":   round(sum(pcts) / n, 1) if n else 0.0,
        "win_rate":  round(wins / n * 100, 1) if n else 0.0,
        "t1_rate":   round(t1_hits / n * 100, 1) if n else 0.0,
        "t2_rate":   round(t2_hits / n * 100, 1) if n else 0.0,
        "days":          len(day_vals),
        "green_days":    green_days,
        "green_day_pct": round(green_days / len(day_vals) * 100, 1) if day_vals else 0.0,
        "best_day_usd":  round(best_day, 2),
        "worst_day_usd": round(min(day_vals), 2) if day_vals else 0.0,
        "concentration": concentration,
        "by_day":        by_day,
    }


def _collect_trades(days: list, send, bot, chat) -> list:
    """
    Stream each day once, and for every qualifying alert grab the option's
    post-fill price path. Returns the raw material the sweep re-scores.
    """
    from preset_backtest import (collect_day, _expiry_to_iso, _fetch_option_daily,
                                 _fetch_option_intraday, INTRADAY_FILL)
    from datetime import timedelta

    trades = []
    for i, date in enumerate(days, 1):
        try:
            day = collect_day(date)
        except Exception as e:
            print(f"[SWEEP] {date}: {e}")
            continue

        for a in day.get("alerts", []):
            flow = float(a.get("price") or 0)
            if flow <= 0:
                continue

            # If a roll applies, the contract we'd actually TAKE is the rolled
            # one — sweep that contract's path and price, not the original's.
            _roll = a.get("roll") or {}
            if _roll.get("available") and _roll.get("price", 0) > 0 and _roll.get("occ"):
                occ     = f"O:{_roll['occ']}"
                flow    = float(_roll["price"])     # the rolled contract's price
                exp_iso = _expiry_to_iso(_roll.get("expiry", ""))
                traded  = "rolled"
            else:
                occ     = f"O:{a['ticker']}{_occ_tail(a)}"
                exp_iso = _expiry_to_iso(a.get("expiry", ""))
                traded  = "original"
            if not exp_iso:
                continue

            # Rebuild the same bar path the single-day backtest uses
            path, day0, used_intraday = [], [], False
            epoch = a.get("timestamp")
            if INTRADAY_FILL and epoch:
                raw = _fetch_option_intraday(occ, date)
                try:
                    ae = float(epoch)
                    day0 = [b for b in raw if b["ts"] >= ae]
                except Exception:
                    day0 = []
                if day0:
                    used_intraday = True
                    path.extend(day0)
            start = date
            if used_intraday:
                try:
                    d0 = datetime.strptime(date, "%Y-%m-%d").date()
                    start = (d0 + timedelta(days=1)).strftime("%Y-%m-%d")
                except Exception:
                    start = date
            path.extend(_fetch_option_daily(occ, start, exp_iso))
            if not path:
                continue

            # Store the FULL path from the alert. The entry discount is applied
            # per-combo in _score_combo, so a deeper discount can legitimately
            # miss the fill — which is exactly what we're trying to measure.
            trades.append({
                "date":       date,
                "ticker":     a["ticker"],
                "flow_price": flow,     # rolled contract's price if a roll applied
                "traded":     traded,
                "path":       path,
            })

        if i % 5 == 0 and i < len(days):
            send(f"… {i}/{len(days)} days streamed ({len(trades)} alerts with data)",
                 bot, chat)

    return trades


def _occ_tail(a: dict) -> str:
    """Rebuild the OCC tail (YYMMDD + C/P + strike) from a result dict."""
    try:
        mm, dd, yy = a["expiry"].split("/")
        cp = "C" if a["direction"] == "call" else "P"
        strike_int = int(round(float(a["strike"]) * 1000))
        return f"{yy}{mm}{dd}{cp}{strike_int:08d}"
    except Exception:
        return ""


def _build_workbook(rows: list, disc_rows: list, trades: list,
                    start: str, end: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    hf    = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F3864")
    ctr   = Alignment(horizontal="center")

    def _hdr(ws, cols):
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hf; cell.fill = hfill; cell.alignment = ctr
        ws.freeze_panes = "A2"

    # ── Tab 1: Entry Discount (the direct answer) ──
    ws_d = wb.active
    ws_d.title = "Entry Discount"
    _hdr(ws_d, ["Entry Disc %", "Alerts", "Fills", "Fill Rate %",
                "$ per ALERT", "$ per FILL", "Total P/L $", "Avg P/L %",
                "Win Rate %", "TP1 Hit %", "TP2 Hit %"])
    for r in sorted(disc_rows, key=lambda x: x["entry_pct"]):
        ws_d.append([r["entry_pct"], r["alerts"], r["fills"], r["fill_rate"],
                     r["usd_per_alert"], r["usd_per_fill"], r["total_usd"],
                     r["avg_pct"], r["win_rate"], r["t1_rate"], r["t2_rate"]])
    # Bold the winner on $/alert
    if disc_rows:
        best_e = max(disc_rows, key=lambda x: x["usd_per_alert"])["entry_pct"]
        for i, r in enumerate(sorted(disc_rows, key=lambda x: x["entry_pct"]), start=2):
            if r["entry_pct"] == best_e:
                for c in range(1, 12):
                    ws_d.cell(row=i, column=c).font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJK", (12, 8, 7, 11, 12, 11, 12, 10, 11, 11, 11)):
        ws_d.column_dimensions[col].width = w

    # ── Tab 2: Full grid, ranked by $/alert ──
    ws = wb.create_sheet("Full Grid")
    _hdr(ws, ["Entry %", "TP1 %", "TP2 %", "Trail %", "Alerts", "Fills",
              "Fill Rate %", "$ per ALERT", "Total P/L $", "Avg P/L %",
              "Win Rate %", "TP1 Hit %", "TP2 Hit %", "Days", "Green Days",
              "Green Day %", "Best Day $", "Worst Day $", "Best-Day Conc %"])
    ranked = sorted(rows, key=lambda x: -x["usd_per_alert"])
    for r in ranked:
        ws.append([r["entry_pct"], r["tp1_pct"], r["tp2_pct"], r["trail_pct"],
                   r["alerts"], r["fills"], r["fill_rate"], r["usd_per_alert"],
                   r["total_usd"], r["avg_pct"], r["win_rate"], r["t1_rate"],
                   r["t2_rate"], r["days"], r["green_days"], r["green_day_pct"],
                   r["best_day_usd"], r["worst_day_usd"], r["concentration"]])
    if ranked:
        for c in range(1, 20):
            ws.cell(row=2, column=c).font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJKLMNOPQRS",
                      (8, 7, 7, 8, 7, 6, 10, 11, 12, 10, 10, 10, 10, 6, 10, 11, 11, 11, 14)):
        ws.column_dimensions[col].width = w

    # ── Tab 3: Top-3 daily (fragility check) ──
    if ranked:
        ws3 = wb.create_sheet("Top3 Daily")
        top3 = ranked[:3]
        _hdr(ws3, ["Date"] + [f"E{r['entry_pct']:.0f}/T1 {r['tp1_pct']:.0f}/"
                              f"T2 {r['tp2_pct']:.0f}/TR {r['trail_pct']:.0f}"
                              for r in top3])
        all_dates = sorted({d for r in top3 for d in r["by_day"]})
        for d in all_dates:
            ws3.append([d] + [r["by_day"].get(d, 0.0) for r in top3])
        ws3.append([])
        ws3.append(["TOTAL"] + [r["total_usd"] for r in top3])
        for c in range(1, len(top3) + 2):
            ws3.cell(row=ws3.max_row, column=c).font = Font(bold=True)
        for col, w in zip("ABCD", (12, 26, 26, 26)):
            ws3.column_dimensions[col].width = w

    # ── Tab 4: the alerts the sweep ran on ──
    ws4 = wb.create_sheet("Alerts")
    _hdr(ws4, ["Date", "Ticker", "Traded", "Contract Price", "Bars"])
    for t in trades:
        ws4.append([t["date"], t["ticker"], t.get("traded", "original"),
                    t["flow_price"], len(t["path"])])
    for col, w in zip("ABCDE", (12, 9, 10, 14, 7)):
        ws4.column_dimensions[col].width = w

    os.makedirs("/tmp", exist_ok=True)
    path = f"/tmp/preset_sweep_{start}_to_{end}.xlsx"
    wb.save(path)
    return path


def _run_sweep_thread(start: str, end: str, bot: str, chat: str):
    from sms import send_telegram, send_telegram_document
    from preset_backtest_range import _weekdays_in_range

    days = _weekdays_in_range(start, end)
    if not days:
        send_telegram(f"No weekdays in {start} → {end}.", bot, chat)
        return

    entries, tp1s, tp2s, trails = _grid()
    combos = len(entries) * len(tp1s) * len(tp2s) * len(trails)

    send_telegram(
        f"🧪 Parameter sweep: {start} → {end}\n"
        f"{len(days)} trading days | up to {combos} combinations\n"
        f"Entry disc: {[f'{int(p*100)}%' for p in entries]}\n"
        f"TP1: {[f'{int(p*100)}%' for p in tp1s]}\n"
        f"TP2: {[f'{int(p*100)}%' for p in tp2s]}\n"
        f"Trail: {[f'{int(p*100)}%' for p in trails]}\n"
        f"Streaming days first (~{len(days)*6} min), then scoring.",
        bot, chat)

    trades = _collect_trades(days, send_telegram, bot, chat)
    if not trades:
        send_telegram("No alerts with option data in that range — nothing to sweep.",
                      bot, chat)
        return

    send_telegram(f"📊 {len(trades)} alerts collected. Scoring…", bot, chat)

    rows = []
    for e in entries:
        for t1 in tp1s:
            for t2 in tp2s:
                if t2 <= t1:
                    continue
                for tr in trails:
                    rows.append(_score_combo(trades, e, t1, t2, tr))

    if not rows:
        send_telegram("No valid combinations (TP2 must exceed TP1).", bot, chat)
        return

    # ── Entry-discount isolation: hold TP/trail at the CURRENT live settings
    #    and vary only the discount, so the comparison is apples-to-apples. ──
    import bullflow_presets as bp
    cur_t1 = bp.TARGET1_PCT
    cur_t2 = bp.TARGET2_PCT
    cur_tr = bp.TRAIL_OFFSET_PCT
    disc_rows = [_score_combo(trades, e, cur_t1, cur_t2, cur_tr) for e in entries]

    best      = max(rows, key=lambda x: x["usd_per_alert"])
    best_disc = max(disc_rows, key=lambda x: x["usd_per_alert"])
    cur = next((r for r in disc_rows
                if abs(r["entry_pct"] - bp.ENTRY_DISCOUNT_PCT * 100) < 0.5), None)

    # Discount comparison table — the direct answer to "is 20% worth it?"
    lines = [f"🧪 Sweep complete: {start} → {end}",
             f"{len(trades)} alerts | {len(rows)} combos",
             "",
             "━━━ ENTRY DISCOUNT (TP/trail at your live settings) ━━━",
             "Disc │ Fill% │ $/alert │ $/fill │ Total $"]
    for r in sorted(disc_rows, key=lambda x: x["entry_pct"]):
        mark = " ←you" if cur and r["entry_pct"] == cur["entry_pct"] else ""
        star = " ⭐" if r["entry_pct"] == best_disc["entry_pct"] else ""
        lines.append(f"{r['entry_pct']:>3.0f}% │ {r['fill_rate']:>4.0f}% │ "
                     f"${r['usd_per_alert']:>+7,.0f} │ ${r['usd_per_fill']:>+6,.0f} │ "
                     f"${r['total_usd']:>+8,.0f}{mark}{star}")

    lines += ["",
              f"⭐ Best discount: {best_disc['entry_pct']:.0f}% "
              f"(${best_disc['usd_per_alert']:+,.0f}/alert, "
              f"{best_disc['fill_rate']:.0f}% fill)"]
    if cur:
        delta = best_disc["usd_per_alert"] - cur["usd_per_alert"]
        lines.append(f"   vs your {cur['entry_pct']:.0f}%: "
                     f"${delta:+,.0f} per alert")

    lines += ["",
              "━━━ BEST OVERALL COMBO ━━━",
              f"🥇 Entry {best['entry_pct']:.0f}% / TP1 {best['tp1_pct']:.0f}% / "
              f"TP2 {best['tp2_pct']:.0f}% / Trail {best['trail_pct']:.0f}%",
              f"   ${best['usd_per_alert']:+,.0f}/alert | ${best['total_usd']:+,.0f} total | "
              f"{best['fill_rate']:.0f}% fill | {best['win_rate']:.0f}% win",
              f"   Green {best['green_days']}/{best['days']} days | "
              f"best day = {best['concentration']:.0f}% of total",
              "",
              "⚠️ Ranked by $/ALERT — this prices in the trades a deeper",
              "discount never fills. Total $ alone would flatter shallow entries.",
              "Full grid + Discount tab in the Excel."]
    send_telegram("\n".join(lines), bot, chat)

    try:
        path = _build_workbook(rows, disc_rows, trades, start, end)
    except Exception as e:
        send_telegram(f"❌ Excel build error: {e}", bot, chat)
        return

    send_telegram_document(path, bot, chat,
                           caption=f"Sweep {start} → {end}: {len(rows)} combos, "
                                   f"{len(trades)} alerts.")
    try:
        os.remove(path)
    except Exception:
        pass


def start_sweep(start: str, end: str, bot: str, chat: str):
    """Validate and launch. Returns (ok, message)."""
    if not (re.match(r'^\d{4}-\d{2}-\d{2}$', start) and re.match(r'^\d{4}-\d{2}-\d{2}$', end)):
        return False, "Both dates must be YYYY-MM-DD. e.g. /preset_sweep 2026-06-01 2026-06-30"
    try:
        d0 = datetime.strptime(start, "%Y-%m-%d").date()
        d1 = datetime.strptime(end,   "%Y-%m-%d").date()
    except Exception:
        return False, "Invalid date(s)."
    if d1 < d0:
        return False, "End date is before start date."
    if (d1 - d0).days + 1 > MAX_RANGE_DAYS:
        return False, f"Range too large. Max {MAX_RANGE_DAYS} days."

    threading.Thread(target=_run_sweep_thread, args=(start, end, bot, chat),
                     daemon=True, name=f"preset_sweep_{start}").start()
    return True, ""
